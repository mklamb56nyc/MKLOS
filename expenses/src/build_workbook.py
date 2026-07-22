"""Build the color-coded review workbook (v4 pattern): GREEN / YELLOW / GRAY tabs.

Tab color = tier. One tab per category, plus a Summary tab with SUM formulas (never
hardcoded totals). After saving, ALWAYS run recalc verification:

    python /mnt/skills/public/xlsx/scripts/recalc.py output/workbook_vN.xlsx
    # (in MKLOS: use LibreOffice headless or open in Excel to confirm zero formula errors)
"""
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

import sys
sys.path.insert(0, str(Path(__file__).parent))
from ingest import safe_sheet_name

ROOT = Path(__file__).resolve().parents[1]
PICKLE = ROOT / "output" / "all_tx_categorized.pkl"

TAB_COLORS = {"GREEN": "00B050", "YELLOW": "FFC000", "GRAY": "808080"}
VERSION = 5  # v4 was the last manual workbook; automation starts at v5


def build(version: int = VERSION) -> Path:
    tx = pd.read_pickle(PICKLE)
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    cols = ["date", "account", "description", "amount", "treatment",
            "split_pct", "deductible_amount", "flag", "rule_note"]
    headers = ["Date", "Account", "Description", "Amount", "Treatment",
               "Split %", "Deductible", "Flag", "Note"]

    summary_rows = []
    for tier in ["GREEN", "YELLOW", "GRAY"]:
        for cat, grp in tx[tx["tier"] == tier].groupby("category"):
            ws = wb.create_sheet(safe_sheet_name(f"{tier[:1]}-{cat}"))
            ws.sheet_properties.tabColor = TAB_COLORS[tier]
            for j, h in enumerate(headers, 1):
                c = ws.cell(1, j, h)
                c.font = Font(name="Arial", bold=True)
            for i, (_, row) in enumerate(grp.sort_values("date").iterrows(), 2):
                for j, col in enumerate(cols, 1):
                    v = row[col]
                    if col == "date":
                        v = v.strftime("%Y-%m-%d")
                    elif pd.isna(v):
                        v = ""
                    ws.cell(i, j, v).font = Font(name="Arial")
            n = len(grp)
            # Formulas, never hardcoded totals
            ws.cell(n + 2, 3, "TOTAL").font = Font(name="Arial", bold=True)
            ws.cell(n + 2, 4, f"=SUM(D2:D{n + 1})").font = Font(name="Arial", bold=True)
            ws.cell(n + 2, 7, f"=SUM(G2:G{n + 1})").font = Font(name="Arial", bold=True)
            summary_rows.append((tier, cat, ws.title, n))

    # Summary tab with cross-sheet SUM formulas (quote sheet names — may contain spaces)
    for j, h in enumerate(["Tier", "Category", "Count", "Spend", "Deductible"], 1):
        summary.cell(1, j, h).font = Font(name="Arial", bold=True)
    for i, (tier, cat, sheet, n) in enumerate(summary_rows, 2):
        summary.cell(i, 1, tier).font = Font(name="Arial")
        summary.cell(i, 2, cat).font = Font(name="Arial")
        summary.cell(i, 3, n).font = Font(name="Arial")
        summary.cell(i, 4, f"=SUM('{sheet}'!D2:D{n + 1})").font = Font(name="Arial")
        summary.cell(i, 5, f"=SUM('{sheet}'!G2:G{n + 1})").font = Font(name="Arial")
    last = len(summary_rows) + 1
    summary.cell(last + 1, 2, "TOTAL").font = Font(name="Arial", bold=True)
    summary.cell(last + 1, 4, f"=SUM(D2:D{last})").font = Font(name="Arial", bold=True)
    summary.cell(last + 1, 5, f"=SUM(E2:E{last})").font = Font(name="Arial", bold=True)

    out = ROOT / "output" / f"expenses_workbook_v{version}.xlsx"
    wb.save(out)
    print(f"Saved {out} — now run recalc verification before treating it as final.")
    return out


if __name__ == "__main__":
    build()
