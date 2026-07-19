import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
NAVY="1F3864"; BLUE="2E5496"; LTBLUE="D6E0F0"; LTBLUE2="EAF0FA"; GREY="F2F2F2"
AMBER="FCE4B8"; AMBER2="F9D9A0"; GREEN="C6EFCE"; GREEN2="D9EAD3"; GREENM="B6D7A8"
WHITE="FFFFFF"; RED="F4CCCC"; FONT="Arial"
thin=Side(style="thin",color="BFBFBF"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
TIERFILL={"Ambitious":RED,"Match":AMBER,"High confidence":GREEN}

def hdr(c,fill=NAVY,color=WHITE,size=11):
    c.font=Font(name=FONT,bold=True,color=color,size=size)
    c.fill=PatternFill("solid",fgColor=fill)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=border
def cell(c,bold=False,fill=None,align="left",valign="top",color="000000",size=9,wrap=True):
    c.font=Font(name=FONT,bold=bold,color=color,size=size)
    if fill: c.fill=PatternFill("solid",fgColor=fill)
    c.alignment=Alignment(horizontal=align,vertical=valign,wrap_text=wrap); c.border=border
def title_row(ws,text,span,height=30,size=14):
    ws.merge_cells(f"A1:{span}1"); t=ws["A1"]; t.value=text
    t.font=Font(name=FONT,bold=True,color=WHITE,size=size)
    t.fill=PatternFill("solid",fgColor=NAVY)
    t.alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[1].height=height

# =========================================================
# TAB 1 — APPLICATION TRACKER (all 21 kept; Tier + shortlist added)
# =========================================================
ws=wb.active; ws.title="Application Tracker"
title_row(ws,"Toby — College Application Tracker  |  2026–27 (Fall 2027)  |  all 21 schools; the 10-school working shortlist is highlighted green","N")
cols=["School","Tier","On\nshortlist","Application system","ED0 / SSEN","ED I / SCEA","EA","ED II","RA / RD",
      "Testing","Recommendation letters (from whom)","Essays required","App fee","Key flags"]
for j,h in enumerate(cols,1):
    hdr(ws.cell(row=2,column=j,value=h), fill=NAVY if j==1 else BLUE)
ws.row_dimensions[2].height=40

# row = [school, tier, shortlist(bool), system, ed0, edI, ea, ed2, rd, testing, req, recs, essays, fee, flags]
R=[
["MIT","Ambitious",True,"MIT portal (not Common App)","—","—","Nov 1\n→ mid-Dec","—","Jan 4\n→ mid-Mar",
 "REQUIRED",True,"2 teachers — one math/science + one humanities/social-science/language; + School Report (counselor)",
 "MIT portal: 4 main essays (~100–200w) + 4 short-response Qs (40–50w) + info box; 4 activities. No Common App essay.","$75",
 "Own system & timeline. EA non-binding, non-restrictive. Midyear via MIT 'Feb Updates' form. Aid: CSS + FAFSA."],
["Yale","Ambitious",True,"Common App / Coalition","—","SCEA Nov 1\n→ mid-Dec","—","—","Jan 2\n→ by Apr 1",
 "REQUIRED*",True,"2 teachers + counselor",
 "Common App PS + academic-interest (≤200w) + Why-Yale (≤125w) + 4 short takes (≤200 char) + one 400w essay (1 of 3).","$80",
 "SCEA is RESTRICTIVE — no other private early plan. *AP/IB no longer replace SAT/ACT."],
["University of Chicago","Ambitious",True,"Common App / Coalition","Sep 1–Oct 15\n→ before Nov 1","ED I ~Nov 1–3*\n→ mid-Dec","~Nov 1–3*\n→ mid-Dec","~Jan 2–6*\n→ mid-Feb","~Jan 2–6*\n→ late Mar",
 "Test-optional",False,"2 teachers + counselor. SSEN: may reuse Summer Session recs; counselor letter waivable.",
 "Common App PS + Why-UChicago (300–600w) + extended 'Uncommon' essay (500–700w).","$75",
 "Toby SSEN-ELIGIBLE via pre-college Intensive Japanese (binding ED0, decision before Nov 1 — see Strategy). EA non-restrictive; ED I/II & SSEN BINDING. No CSS (FAFSA + UChicago worksheet). *Confirm dates."],
["Brown","Match",False,"Common App only","—","ED Nov 1\n→ mid-Dec","—","—","Jan 5\n→ early Apr",
 "REQUIRED",True,"2 teachers + counselor","Common App PS + 3 essays (200–250w) + 3 short answers","$80","ED is BINDING."],
["Columbia","Match",True,"Common App / Coalition","—","ED Nov 1\n→ mid-Dec","—","—","Jan 1\n→ late Mar",
 "Test-optional this cycle",False,"2 teachers + counselor","Common App PS + list Q (100w) + 5 short essays (150w each)","$85",
 "ED BINDING. Test-required starting NEXT cycle. Apply to Columbia College or Engineering."],
["Cornell","Match",False,"Common App","—","ED Nov 1\n→ mid-Dec","—","—","Jan 2\n→ late Mar/Apr",
 "REQUIRED",True,"2 teachers + counselor","Common App PS + university essay (~350w) + college-specific essay (Eng = 7 pieces)","$85",
 "ED BINDING. Apply to ONE specific undergraduate college; prompts differ by college."],
["Princeton","Ambitious",False,"Common App","—","—","SCEA Nov 1\n→ mid-Dec","—","Jan 1\n→ late Mar",
 "Test-optional this cycle",False,"2 teachers (core) + counselor",
 "Common App PS + academic + activity + civic-engagement essays + 3 short answers (50w) + GRADED PAPER","$75",
 "SCEA RESTRICTIVE (like Yale). Needs a graded English/history paper."],
["UPenn","Match",True,"Common App / Coalition","—","ED Nov 1\n→ mid-Dec","—","—","Jan 5\n→ by Apr 1",
 "REQUIRED",True,"2 teachers + counselor","Common App PS + thank-you note + community essay + school-specific essay (all 150–200w)","$75",
 "ED BINDING. Apply to one of four schools (A&S, Wharton, Engineering, Nursing)."],
["Williams","Match",False,"Common App / Coalition","—","ED Nov 15\n→ by Dec 15","—","—","Jan 5\n→ early Apr",
 "Test-optional",False,"2 teachers + counselor","Common App PS only (optional academic paper)","$65","ED BINDING. Later Nov 15 ED date."],
["Swarthmore","Match",True,"Common App / Coalition","—","ED I Nov 15\n→ mid-Dec","—","Jan 4\n→ mid-Feb","Jan 4\n→ by Apr 1",
 "Test-optional",False,"2 teachers + counselor","Common App PS + 2 short essays (250w each: identity/community + curiosity)","$60","ED I & II BINDING."],
["Wesleyan","High confidence",True,"Common App / Coalition","—","ED I Nov 15\n→ by Dec 15","—","Jan 1\n→ by Feb 15","Jan 1\n→ late Mar",
 "Test-optional",False,"2 teachers (1 STEM + 1 humanities pref.) + counselor",
 "Common App PS only (optional creative supplement)","$65","ED I & II BINDING."],
["Vassar","High confidence",True,"Common App / Coalition","—","ED I Nov 15\n→ mid-Dec","—","~Jan 14*\n→ mid-Feb","~Jan 1–4*\n→ late Mar",
 "Test-optional",False,"2 teachers + counselor","Common App PS + 1 required essay (300w, choose 1 of 2) + optional 'Your Space'","$65",
 "ED I & II BINDING. *Confirm ED II/RD dates."],
["Oberlin","High confidence",False,"Common App / Coalition","—","ED I ~Nov 15\n→ Dec","—","~Jan 2\n→ Feb","Jan 15*\n→ late Mar",
 "Test-optional",False,"1 teacher (core) + counselor","Common App PS only. No required supplement.","~$65",
 "ED I & II BINDING. *Also references an EA option — confirm."],
["Connecticut College","High confidence",False,"Common App","—","ED I Nov 15\n→ mid-Dec","confirm date*","Jan 15\n→ Feb","Jan 15\n→ late Mar",
 "Test-optional",False,"2 teachers + counselor (+ optional peer rec)","Common App PS only. No required supplement.","$0 (no fee)",
 "ED I/II BINDING. No application fee. *Confirm EA deadline."],
["Tufts","Match",False,"Common App / Coalition","—","ED I ~Nov 1–3\n→ mid-Dec","—","~Jan 4–5\n→ early Feb","~Jan 4–5\n→ by Apr 1",
 "Test-optional*",False,"Counselor + ≥1 core teacher","Common App PS + 2 short supplements (incl. program-specific)","$75","ED I & II BINDING. *Confirm TO extension."],
["NYU","Match",False,"Common App","—","ED I Nov 1\n→ mid-Dec","—","Jan 1\n→ mid-Feb","Jan 5\n→ ~Apr 1",
 "Test-optional / flexible",False,"1 teacher + counselor","Common App PS + 'Why NYU' (~400w)","$80","ED I & II BINDING. Apply to a specific NYU school."],
["University of Washington","High confidence",False,"ApplyWeb (UW) or Common App","—","—","—","—","Nov 15 (only)\n→ by mid-Mar",
 "Test-optional",False,"None required","UW-specific essay (~650w) + short response (~300w). Does NOT use the Common App essay.","$80",
 "One deadline for all (public). Direct-to-major for some majors (e.g. CS/eng)."],
["RIT","High confidence",True,"Common App or RIT app","—","ED I Nov 1\n→ ~early Jan","Nov 1\n→ late Mar","Jan 1\n→ ~early Feb","Jan 15\n→ late Mar",
 "Test-optional",False,"1 recommendation","Common App PS + 1 RIT supplemental essay","$65",
 "ED I/II BINDING; EA non-binding. STEM + ACT → add optional Science section."],
["Ithaca College","High confidence",False,"Common App / Coalition","—","ED Nov 1\n→ by Dec 15","Dec 1\n→ rolling by Feb 1","—","Feb 1\n→ rolling by Apr 15",
 "Test-optional",False,"1 recommendation","Common App PS. No general supplement (arts: audition/portfolio).","$60 (free ED)",
 "ED BINDING. Toby well above profile — likely admit / safety."],
["Baruch (CUNY)","High confidence",False,"CUNY app (1 app → up to 6 colleges)","—","—","—","—","Feb 1 priority\n→ rolling from mid-Feb",
 "Test-optional (thru Spr 2027)",False,"1 academic recommendation (Macaulay needs 2)",
 "CUNY application essay (some colleges have own topics)","~$65 (one fee)",
 "~48% overall but SAFETY for his stats (median SAT ~1280). Zicklin business more competitive (~40–45%). Macaulay Honors = separate app (2 essays + 2 recs)."],
["University of Minnesota","High confidence",True,"Common App or Golden Gopher app","—","—","EA I: Nov 1\nEA II: Dec 1\n→ by ~Jan 31","—","RD: Jan 1\n→ rolling",
 "Test-optional",False,"None required (self-reported grades; recs & transcript not required)",
 "1 short essay (which of UMN's 8 colleges & 150+ majors, and why).","$55",
 "Twin Cities flagship SAFETY (~75–80%). Non-binding EA. No demonstrated interest. Auto-considered for other UMN colleges."],
]
r=3
for row in R:
    school,tier,short,system,ed0,edI,ea,ed2,rd,testing,req,recs,essays,fee,flags=row
    mark="✓" if short else ""
    vals=[school,tier,mark,system,ed0,edI,ea,ed2,rd,testing,recs,essays,fee,flags]
    if short: base=GREEN2
    else: base=LTBLUE2 if r%2 else WHITE
    for j,v in enumerate(vals,1):
        c=ws.cell(row=r,column=j,value=v)
        if j==1: cell(c,bold=True,fill=(GREENM if short else LTBLUE),valign="center")
        elif j==2: cell(c,fill=TIERFILL.get(tier,base),align="center",valign="center",bold=True,size=8)
        elif j==3: cell(c,fill=base,align="center",valign="center",bold=True,color="38761D")
        elif j in (5,6,7,8,9): cell(c,fill=base,align="center",valign="center")
        elif j==10: cell(c,fill=(AMBER if req else GREEN),align="center",valign="center",bold=req,size=8)
        elif j==13: cell(c,fill=base,align="center",valign="center")
        else: cell(c,fill=base)
    longest=max(len(edI),len(recs),len(essays),len(flags))
    ws.row_dimensions[r].height=max(46, min(150, 14*(1+longest//28)))
    r+=1
widths={"A":18,"B":13,"C":6,"D":16,"E":14,"F":14,"G":13,"H":12,"I":15,"J":13,"K":24,"L":26,"M":9,"N":30}
for k,v in widths.items(): ws.column_dimensions[k].width=v
ws.freeze_panes="D3"
ws.merge_cells(f"A{r}:N{r}")
fn=ws.cell(row=r,column=1,value="GREEN rows = the 10-school working shortlist (Ambitious: MIT, Yale, UChicago · Match: UPenn, Columbia, Swarthmore · High confidence: Vassar, RIT, Wesleyan, U Minnesota). The Deadline Calendar, Task List, and Strategy tabs are scoped to these 10; this grid keeps all 21 for reference. Tier colours: red = Ambitious, amber = Match, green = High confidence. Each round cell = deadline → notification. Amber Testing = SAT/ACT required.")
fn.font=Font(name=FONT,italic=True,size=8,color="808080")
fn.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
ws.row_dimensions[r].height=54
print("tab1 done, rows through", r)

# =========================================================
# TAB 2 — ESSAY PROMPTS (all schools kept as reference; U Minn fixed)
# =========================================================
wse=wb.create_sheet("Essay Prompts")
title_row(wse,"Actual Essay Prompts — official sources (all schools; shortlisted 10 marked ★)","D")
for j,h in enumerate(["Essay piece","Prompt / question","Limit","Required?"],1):
    hdr(wse.cell(row=2,column=j,value=h),fill=BLUE)
wse.row_dimensions[2].height=20
def banner(row,text):
    wse.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
    c=wse.cell(row=row,column=1,value=text)
    c.font=Font(name=FONT,bold=True,color=WHITE,size=10)
    c.fill=PatternFill("solid",fgColor=NAVY)
    c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
    wse.row_dimensions[row].height=28
essay_data=[
 ("B","★ MIT  —  MIT portal, 2026–27. No Common App essay.  Link: mitadmissions.org/apply/firstyear/essays-activities-academics"),
 ("R","Main essay 1 — Academic interest","Pick the field of study that most appeals to you now (drop-down), then reflect on what led you to that interest.","~100–200 words","Required"),
 ("R","Main essay 2 — Your own path","Some follow well-trodden paths while others blaze their own: describe a way you've done something different from what was expected in your educational journey.","~100–200 words","Required"),
 ("R","Main essay 3 — Problems & people","Reflect on how your personal and academic experiences shape the problems you'd want to tackle with an MIT education — and who you'd work on them with.","~100–200 words","Required"),
 ("R","Main essay 4 — Unexpected challenge","How did you handle a situation or challenge you didn't expect, and what did you learn?","~100–200 words","Required"),
 ("R","Additional info (open box)","Open space to share anything else MIT should know (MIT expects most students to write something).","Open-ended","Optional (encouraged)"),
 ("R","Short response — Fun","What do you do just for fun?","40–50 words","Required"),
 ("R","Short response — Admire","Someone you admire (in person or from afar) — and why.","40–50 words","Required"),
 ("R","Short response — Talk for hours","A topic (academic or not) you could talk about for hours.","40–50 words","Required"),
 ("R","Short response — Generalist/specialist","MIT values both generalists and specialists — which best describes you, and why?","40–50 words","Required"),
 ("R","Activities","List the four activities that mean the most to you.","Up to 4","Required"),

 ("B","★ YALE  —  Common App / Coalition / QuestBridge + Yale questions. Most recent confirmed set (2025–26).  Link: admissions.yale.edu/essay-topics"),
 ("R","Common App personal statement","Standard 650-word Common App essay.","≤650 words","Required"),
 ("R","Academic areas (selection)","Select up to three academic areas that fit your interests/goals (selection, no writing).","Pick ≤3","Required"),
 ("R","Academic-interest essay","A topic or idea that excites you within those areas — why are you drawn to it?","≤200 words","Required"),
 ("R","Why Yale","Reflect on how your interests, values, and/or experiences have drawn you to Yale.","≤125 words","Required"),
 ("R","Short take — Inspires","What inspires you?","≤200 characters","Required (Common/Coalition)"),
 ("R","Short take — Teach/write/create","If you could teach any college course, write a book, or create any original piece of art — what would it be?","≤200 characters","Required (Common/Coalition)"),
 ("R","Short take — Influence","Other than a family member, who has significantly influenced you — and what was the impact?","≤200 characters","Required (Common/Coalition)"),
 ("R","Short take — Not elsewhere","Something about you not included anywhere else in your application.","≤200 characters","Required (Common/Coalition)"),
 ("R","400-word essay (choose 1 of 3)","A: an issue you discussed with someone holding an opposing view — why meaningful.  B: a community you feel connected to — why meaningful.  C: an element of your experience that will enrich your college — how it shaped you.","≤400 words","Required (choose 1)"),
 ("R","Arts / STEM supplement","Optional Visual Art, Dance, Music, Film, or STEM research materials.","Portfolio","Optional"),

 ("B","★ UNIVERSITY OF CHICAGO  —  Common App / Coalition + UChicago questions, 2026–27. Five 'Uncommon' prompts change yearly — exact wording at link.  Link: collegeadmissions.uchicago.edu"),
 ("R","Why UChicago","How does UChicago, as you know it now, satisfy your desire for a particular kind of learning, community, and future? Be specific.","~300–600 words","Required"),
 ("R","Extended 'Uncommon' essay","Choose ONE of UChicago's five signature quirky prompts (change yearly), OR invent your own / use any past prompt.","~500–700 words","Required"),

 ("B","BROWN  —  Common App. 2025–26 set.  Link: admission.brown.edu/apply/how-apply"),
 ("R","Open Curriculum / academic interest","Tell us about academic interests that excite you and how you might pursue them at Brown (Open Curriculum).","200–250 words","Required"),
 ("R","Where you grew up","How an aspect of your growing up inspired or challenged you, and what you'd contribute to Brown.","200–250 words","Required"),
 ("R","Joy","Something that brings you joy.","200–250 words","Required"),
 ("R","Short answers (×3)","Three words that describe you; a ~100-word short answer; a one-sentence Why Brown (≤50 words).","3w / ~100w / ≤50w","Required"),

 ("B","★ COLUMBIA  —  Common App / Coalition. 2025–26 set.  Link: undergrad.admissions.columbia.edu/apply/firstyear"),
 ("R","List question","List texts/resources/outlets that shaped your intellectual development outside coursework (list only; no author names or explanations).","≤100 words","Required"),
 ("R","Lived experience","An aspect of your life/lived experience important to you, and how it shaped how you'd learn from and contribute to Columbia.","≤150 words","Required"),
 ("R","Perspectives","How you learn and thrive across a wide range of perspectives.","≤150 words","Required"),
 ("R","Resilience / adversity","A time you faced adversity and what you took from it.","≤150 words","Required"),
 ("R","Why Columbia","Why you're drawn to Columbia specifically.","≤150 words","Required"),
 ("R","Why major","Why your intended field (differs for Columbia College vs. Engineering).","≤150 words","Required"),

 ("B","CORNELL  —  Common App. Apply to ONE college; each has its own essay. 2026–27 university-essay wording posts late summer.  Link: admissions.cornell.edu/how-to-apply/first-year-applicants/cornell-first-year-writing-supplement-prompts"),
 ("R","Cornell University essay","How you've been shaped by one of the communities you belong to (define community however you like).","~350 words","Required (everyone)"),
 ("R","College-specific essay","One essay for the Cornell college you apply to (CALS ~500w; A&S, AAP, Brooks, Dyson, ILR ~650w; Human Ecology ~600w).","500–650 words","Required"),
 ("R","Engineering (if applying)","Two 200-word essays + four 100-word short answers (seven pieces).","2×200w + 4×100w","Engineering only"),

 ("B","PRINCETON  —  Common App + Princeton questions. 2025–26 set.  Link: admission.princeton.edu/apply"),
 ("R","Academic interest (Why major)","A.B./undecided: which academic areas most pique your curiosity, and how do Princeton's programs suit them?  B.S.E.: why study engineering at Princeton.","~250 words","Required"),
 ("R","Activity / extracurricular","Briefly elaborate on a meaningful activity, organization, work experience, or hobby.","~150 words","Required"),
 ("R","Service & civic engagement","Princeton's commitment is to service & civic engagement. How does your own story intersect with these ideals?","≤250 words","Required"),
 ("R","Short answers (×3)","A new skill you'd like to learn in college; what brings you joy; a song that's the soundtrack of your life.","≤50 words each","Required"),
 ("R","Graded written paper","Submit a graded paper (preferably English/history) with the teacher's marks — a required document.","1–2 pages","Required"),

 ("B","★ UPENN  —  Common App / Coalition. Apply to one of four schools. 2025–26 set.  Link: admissions.upenn.edu/how-to-apply/first-year-applicants"),
 ("R","Thank-you note","Write a short thank-you note to someone you haven't yet thanked (you're encouraged to actually share it).","150–200 words","Required"),
 ("R","Community at Penn","How will you explore community at Penn? How will Penn shape your perspective, and how will you shape Penn?","150–200 words","Required"),
 ("R","School-specific essay","One essay for your intended school (Arts & Sciences, Wharton, Engineering, or Nursing).","150–200 words","Required"),
 ("R","Dual-degree / specialized (if applying)","M&T, Huntsman, LSM, VIPER, NHCM, DMD, etc. require extra program essays.","250–650 words","If applicable"),

 ("B","★ SWARTHMORE  —  Common App / Coalition. 2025–26 set.  Link: swarthmore.edu/admissions-aid/application-materials-deadlines"),
 ("R","Identity & community","Tell us about a part of your identity and how you engage with communities (diverse, equitable, inclusive inquiry).","≤250 words","Required"),
 ("R","Curiosity","A topic or idea that recently sparked your curiosity.","≤250 words","Required"),

 ("B","★ VASSAR  —  Common App / Coalition. 2025–26 set.  Link: vassar.edu/admission/apply/requirements"),
 ("R","Community / identity (choose 1 of 2)","Option A (engaged pluralism) or Option B (a community that shaped you): an important part of your identity/background and how it shapes what you'd bring to Vassar.","≤300 words","Required (choose 1)"),
 ("R","'Your Space' (optional)","Optional creative upload — poetry, art, photography, video, collage, etc.","Any medium","Optional (encouraged)"),

 ("B","TUFTS  —  Common App / Coalition. 2025–26 set.  Link: admissions.tufts.edu/apply/applying-to-tufts"),
 ("R","How you engaged with Tufts","How you've learned about and engaged with Tufts during your search (demonstrated interest counts).","~150 words","Required"),
 ("R","Second short answer (by program)","A&S / Engineering: choose one (intellectual curiosity; how your environment shaped you; your journey with social justice). SMFA/Combined: portfolio-linked.","~200–250 words","Required"),

 ("B","NYU  —  Common App. 2025–26 set.  Link: nyu.edu/admissions"),
 ("R","Why NYU","What motivated you to apply to NYU, and why the particular campus/school/college/program/area of study?","~400 words","Required"),

 ("B","UNIVERSITY OF WASHINGTON  —  UW's own app or Common App. UW does NOT read the Common App essay.  Link: admit.washington.edu/apply/first-year/how-to-apply"),
 ("R","Personal statement","A story from your life that demonstrates or shaped your character (UW's personal statement, used instead of the Common App essay).","up to ~650 words","Required"),
 ("R","Short response","Why are you excited to push your education outside the areas you're most comfortable with?","~300 words","Required*"),
 ("R","Additional information (optional)","Optional space for hardships, work/family obligations, or unusual limitations/opportunities.","~200 words","Optional"),

 ("B","★ RIT  —  Common App or RIT application.  Link: rit.edu/admissions/first-year-application"),
 ("R","RIT supplemental essay","One RIT-specific essay on your intended path — the outcome you hope to achieve and how you've prepared for RIT's hands-on education.","1 essay","Required"),

 ("B","BARUCH (CUNY)  —  one CUNY application (up to 6 colleges). Essay + 1 rec now required.  Link: cuny.edu/admissions/undergraduate/apply"),
 ("R","CUNY application essay","A personal essay is now required. Some colleges use their own topics (e.g., Hunter: 'tell us something meaningful about yourself not reflected elsewhere', ≤500w).","~500 words","Required"),
 ("R","Macaulay Honors (if applying)","Macaulay Honors College requires two additional essays (separate application).","2 essays","Macaulay only"),

 ("B","★ UNIVERSITY OF MINNESOTA (Twin Cities)  —  Common App or Golden Gopher app. No recs/transcript required (self-reported grades).  Link: admissions.tc.umn.edu/apply"),
 ("R","UMN short essay","The U of M has 8 freshman-admitting colleges and 150+ majors — which college(s)/major(s) fit your interests, and why?","~1 short essay","Required"),

 ("B","NO SCHOOL-SPECIFIC ESSAY  —  Common App personal statement (≤650 words) only:"),
 ("R","Williams","Personal statement only. Optional 3–5 page graded/academic paper.  Link: williams.edu/admission-aid/how-to-apply/first-year","≤650 words","PS only"),
 ("R","Oberlin","Personal statement only; no supplement.  Link: oberlin.edu/admissions-and-aid/arts-and-sciences/first-year-applicants","≤650 words","PS only"),
 ("R","Connecticut College","Personal statement only; no supplement and no application fee.  Link: conncoll.edu/admission/apply","≤650 words","PS only"),
 ("R","Ithaca College","Personal statement; no general supplement (some arts majors audition/portfolio).  Link: ithaca.edu/admission/undergraduate-admission/apply-ic","≤650 words","PS only"),
]
r=3
for item in essay_data:
    if item[0]=="B":
        banner(r,item[1]); r+=1
    else:
        _,piece,prompt,limit,req=item
        base=LTBLUE2 if r%2 else WHITE
        cell(wse.cell(row=r,column=1,value=piece),bold=True,fill=LTBLUE)
        cell(wse.cell(row=r,column=2,value=prompt),fill=base)
        cell(wse.cell(row=r,column=3,value=limit),fill=base,align="center",valign="center")
        cell(wse.cell(row=r,column=4,value=req),fill=base,align="center",valign="center")
        wse.row_dimensions[r].height=max(28,14*(1+len(prompt)//72))
        r+=1
for k,v in {"A":24,"B":78,"C":18,"D":18}.items(): wse.column_dimensions[k].width=v
wse.freeze_panes="A3"
wse.merge_cells(f"A{r}:D{r}")
fnn=wse.cell(row=r,column=1,value="★ = on the 10-school shortlist. Prompts shown are the current published questions (mostly 2025–26; MIT & UChicago are 2026–27). The official link in each banner has the exact, current verbatim wording — 2026–27 supplements finalize when the Common App opens ~Aug 1. UChicago's five creative 'Uncommon' prompts change yearly.")
fnn.font=Font(name=FONT,italic=True,size=8,color="808080")
fnn.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
wse.row_dimensions[r].height=44
print("tab2 done")

# =========================================================
# TAB 3 — DEADLINE CALENDAR (scoped to the 10-school shortlist)
# =========================================================
ws2=wb.create_sheet("Deadline Calendar")
title_row(ws2,"Deadline Calendar — shortlist of 10 (2026–27)","D")
for j,h in enumerate(["Date","What's due","Schools / round","Notes"],1):
    hdr(ws2.cell(row=2,column=j,value=h),fill=BLUE)
ws2.row_dimensions[2].height=20
cal=[
["Aug 1, 2026","Common App & most portals open","Yale, UChicago, UPenn, Columbia, Swarthmore, Vassar, Wesleyan, RIT","MIT portal ~Aug/Sept; U Minnesota via Common App or Golden Gopher."],
["Sept 1 – Oct 15, 2026","UChicago SSEN (ED0) window","University of Chicago (binding)","Toby's earliest, binding shot (pre-college students). Decision BEFORE Nov 1; relaxed rec rules."],
["Oct 1, 2026","FAFSA (2027–28) & CSS Profile open","All (aid)","CSS for Yale, MIT, UPenn, Columbia, Swarthmore, Vassar, Wesleyan, RIT. UChicago = own worksheet. U Minnesota = FAFSA only."],
["Nov 1, 2026","Biggest early deadline","MIT EA · UChicago EA/ED I · UPenn ED · Columbia ED · RIT ED I/EA · U Minnesota EA I","Pick ONE early lane first (Strategy tab). Yale SCEA is also Nov 1 if you go that way."],
["Nov 15, 2026","LAC early deadline","Swarthmore ED I · Vassar ED I · Wesleyan ED I","ED is BINDING — only one allowed."],
["Dec 1, 2026","Early Action II / priority","U Minnesota EA II","Non-binding; apply early for scholarships/housing."],
["Mid-Dec 2026","Early decisions released","MIT, UChicago, UPenn, Columbia, Swarthmore, Vassar, Wesleyan","Admitted to a BINDING ED → enroll & withdraw the rest."],
["Jan 1, 2027","RD / ED II cluster","Columbia RD · Wesleyan ED II/RD · U Minnesota RD","—"],
["Jan 2, 2027","RD","Yale RD","—"],
["Jan 4–5, 2027","RD / ED II cluster","MIT RA (Jan 4) · UChicago RD · Swarthmore ED II/RD (Jan 4) · Vassar RD · UPenn RD (Jan 5)","Confirm exact dates per school."],
["~Jan 14, 2027","ED II","Vassar ED II","Confirm on Vassar's site."],
["Jan 15, 2027","RD","RIT RD","—"],
["~Jan 31, 2027","Early Action I decisions","U Minnesota (EA I)","Rolling thereafter."],
["~Feb 15, 2027","Regular-round aid; MIT Feb Updates","MIT (Feb Updates + midyear) · CSS/FAFSA priority","Counselor sends midyear reports to all."],
["Mar–Apr 2027","Regular decisions released","All 10","MIT mid-Mar · Yale by Apr 1 · others late Mar–early Apr."],
["May 1, 2027","Enrollment deposit due","Wherever admitted","Compare offers & aid, then commit."],
]
r=3
for d in cal:
    date,due,sch,note=d
    base=LTBLUE2 if r%2 else WHITE
    fill = AMBER if ("Nov 1" in date or "Nov 15" in date) else base
    cell(ws2.cell(row=r,column=1,value=date),bold=True,fill=fill,valign="center")
    cell(ws2.cell(row=r,column=2,value=due),fill=base,bold=("Biggest" in due))
    cell(ws2.cell(row=r,column=3,value=sch),fill=base)
    cell(ws2.cell(row=r,column=4,value=note),fill=base)
    ws2.row_dimensions[r].height=max(30,14*(1+len(sch)//48))
    r+=1
for k,v in {"A":16,"B":26,"C":52,"D":46}.items(): ws2.column_dimensions[k].width=v
ws2.freeze_panes="A3"

# =========================================================
# TAB 4 — PRIORITIZED TASK LIST (scoped to the 10)
# =========================================================
ws3=wb.create_sheet("Prioritized Task List")
title_row(ws3,"Prioritized Task List — shortlist of 10 · do these in order","G")
ws3.merge_cells("A2:E2")
p=ws3.cell(row=2,column=1,value="Tasks complete:")
p.font=Font(name=FONT,bold=True,size=10); p.alignment=Alignment(horizontal="right",vertical="center")
for j,h in enumerate(["#","Task","School(s)","Category","Target window","Hard deadline","Status"],1):
    hdr(ws3.cell(row=3,column=j,value=h),fill=BLUE)
ws3.row_dimensions[3].height=20
tasks=[
("Decide the EARLY LANE (this gates everything). With this shortlist only Yale is restrictive — see Strategy tab.","All","Strategy","Now–Aug 15","Before writing early essays"),
("Confirm test scores ready to send; decide whether to retake SAT (1590 already) — recommend NOT.","All","Testing","July","Aug SAT reg ~Aug 7 if retaking"),
("If UChicago is a top choice: prepare & submit the SSEN (ED0) app (Sept 1–Oct 15). Reuse Summer Session teacher recs; counselor letter waivable. Binding, but you hear BEFORE Nov 1.","UChicago","Strategy","Sept 1–Oct 15","Oct 15, 2026"),
("Line up recommenders: counselor + 1 math/science + 1 humanities/language teacher (covers MIT's split & every 2-rec school on the list).","All","Recs","Aug–early Sept","Well before Nov 1"),
("Finalize the Common App personal statement (650w) — draft already underway.","Most","Essays","August","Before early deadline"),
("Create accounts: Common App, MIT portal, RIT (or Common App), U Minnesota (Common App or Golden Gopher).","All","Setup","Early Sept","—"),
("Complete Common App core (profile, activities, honors).","Common App schools","Application","September","—"),
("Write MIT's 4 main essays + 4 short responses (separate system).","MIT","Essays","Sept","Nov 1 if EA"),
("Draft the early-round supplements first: UChicago (Why + Extended), UPenn (thank-you + community + school), Columbia (list + 5 shorts).","UChicago, UPenn, Columbia","Essays","Sept–Oct","Nov 1"),
("Draft Swarthmore (2), Vassar (1), Wesleyan (creative optional), RIT (1), U Minnesota (1 short) essays.","Swat, Vassar, Wesleyan, RIT, UMN","Essays","Sept–Oct","Nov 1 / Nov 15"),
("Give recommenders a resume/brag sheet; confirm Nov 1 submission.","2-rec schools","Recs","Sept","Nov 1"),
("Complete FAFSA (2027–28) at open (~Oct 1).","All (aid)","Financial aid","October","Early: Nov 1–15"),
("Complete CSS Profile (Yale, MIT, UPenn, Columbia, Swarthmore, Vassar, Wesleyan, RIT). UChicago = own worksheet; U Minnesota = FAFSA only.","Privates","Financial aid","October","Early priority"),
("Confirm counselor sends School Report + transcript where required (not needed for U Minnesota).","All","Recs","October","Nov 1"),
("Finalize & proofread ALL early-round essays.","Per lane","Essays","Mid–late Oct","Nov 1 / Nov 15"),
("SUBMIT Nov 1: your early lane + MIT EA / UChicago (per lane) + Columbia/UPenn (if ED) + RIT (ED I/EA) + U Minnesota EA I.","Nov 1 schools","Submit","By Nov 1","Nov 1, 2026"),
("SUBMIT Nov 15: Swarthmore ED I / Vassar ED I / Wesleyan ED I (if applying early).","Nov 15 schools","Submit","By Nov 15","Nov 15, 2026"),
("Submit U Minnesota EA II if not already done.","U Minnesota","Submit","Late Nov","Dec 1, 2026"),
("Verify all components received (recs, reports, scores) in each portal.","All","Submit","Nov–Dec","—"),
("Mid-Dec: review early results. If admitted to a binding ED, enroll & withdraw the rest; otherwise push remaining schools to RD.","All","Decision","Mid-Dec","—"),
("Finalize & submit Jan apps: Yale (Jan 2), MIT RA (Jan 4), UChicago RD, Columbia RD (Jan 1), UPenn RD (Jan 5), Swarthmore RD (Jan 4), Vassar RD, Wesleyan RD (Jan 1), U Minnesota RD (Jan 1).","Jan cluster","Submit","Late Dec","Jan 1–5, 2027"),
("Submit RIT RD (Jan 15) if not applied early.","RIT","Submit","Early Jan","Jan 15, 2027"),
("Submit MIT 'February Updates & Notes' form + midyear; confirm counselor sends midyear reports to all.","MIT + all","Follow-up","Mid-Feb","Mid-Feb 2027"),
("Complete regular-round financial aid (CSS/FAFSA) — due ~Feb 15.","Privates","Financial aid","February","~Feb 15"),
("Review all decisions (Mar–Apr); compare offers & aid; submit deposit.","All","Decision","Mar–Apr","May 1, 2027"),
]
cat_fill={"Strategy":AMBER2,"Testing":GREY,"Recs":LTBLUE2,"Essays":LTBLUE2,"Setup":GREY,
          "Application":GREY,"Financial aid":GREEN,"Submit":AMBER,"Decision":LTBLUE,"Follow-up":GREY}
start=4; r=start
for i,(task,sch,cat,win,hard) in enumerate(tasks,1):
    cell(ws3.cell(row=r,column=1,value=i),bold=True,align="center",valign="center")
    cell(ws3.cell(row=r,column=2,value=task))
    cell(ws3.cell(row=r,column=3,value=sch),align="center",valign="center")
    cell(ws3.cell(row=r,column=4,value=cat),fill=cat_fill.get(cat),align="center",valign="center",bold=(cat in("Strategy","Submit")))
    cell(ws3.cell(row=r,column=5,value=win),align="center",valign="center")
    cell(ws3.cell(row=r,column=6,value=hard),align="center",valign="center",bold=True)
    cell(ws3.cell(row=r,column=7,value="Not started"),align="center",valign="center")
    ws3.row_dimensions[r].height=max(28,14*(1+len(task)//60))
    r+=1
end=r-1
dv=DataValidation(type="list",formula1='"Not started,In progress,Done,N/A"',allow_blank=True)
ws3.add_data_validation(dv); dv.add(f"G{start}:G{end}")
pv=ws3.cell(row=2,column=6); pv.value=f'=COUNTIF(G{start}:G{end},"Done")&" / {len(tasks)}"'
pv.font=Font(name=FONT,bold=True,size=10,color=NAVY); pv.alignment=Alignment(horizontal="left",vertical="center")
for k,v in {"A":4,"B":66,"C":18,"D":14,"E":14,"F":18,"G":13}.items(): ws3.column_dimensions[k].width=v
ws3.freeze_panes="A4"
print("tab3+4 done")

# =========================================================
# TAB 5 — STRATEGY & FLAGS (scoped to the 10)
# =========================================================
ws4=wb.create_sheet("Strategy & Flags")
ws4.column_dimensions["A"].width=3; ws4.column_dimensions["B"].width=112
def blk(row,text,kind="body"):
    ws4.merge_cells(start_row=row,start_column=2,end_row=row,end_column=2)
    c=ws4.cell(row=row,column=2,value=text)
    if kind=="title":
        c.font=Font(name=FONT,bold=True,size=14,color=WHITE); c.fill=PatternFill("solid",fgColor=NAVY)
        c.alignment=Alignment(horizontal="left",vertical="center"); ws4.row_dimensions[row].height=30
    elif kind=="h":
        c.font=Font(name=FONT,bold=True,size=12,color=NAVY)
        c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True); ws4.row_dimensions[row].height=22
    elif kind=="flag":
        c.font=Font(name=FONT,bold=True,size=10,color="9C4A00"); c.fill=PatternFill("solid",fgColor=AMBER)
        c.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
    else:
        c.font=Font(name=FONT,size=10,color="000000")
        c.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
def h6(row,n): ws4.row_dimensions[row].height=n
R=1
blk(R,"Strategy & Key Flags — working shortlist of 10"); R+=2
blk(R,"Shortlist: Ambitious — MIT, Yale, UChicago.  Match — UPenn, Columbia, Swarthmore.  High confidence — Vassar, RIT, "
      "Wesleyan, University of Minnesota.  (All 21 stay in the Application Tracker for reference.)","body"); h6(R,30); R+=2

blk(R,"1. Pick your EARLY LANE first — it's the decision everything hangs on","h"); R+=1
blk(R,"On this shortlist, only ONE school has a restrictive early plan: Yale (SCEA). The rest of the early options are binding "
      "EDs or non-binding EAs. So Toby chooses one of three lanes:","body"); h6(R,42); R+=1
blk(R,"   Lane A — Yale SCEA (restrictive). Rules out EVERY other private early plan: no MIT EA, no UChicago EA, and no ED "
      "(UPenn, Columbia, Swarthmore, Vassar, Wesleyan, RIT). Everything else goes Regular. U Minnesota EA (public) is still allowed.","flag"); h6(R,44); R+=1
blk(R,"   Lane B — One binding ED at a single favourite (UChicago via SSEN or ED I, UPenn, Columbia, Swarthmore, Vassar, "
      "Wesleyan, or RIT). You may STILL add non-binding MIT EA + UChicago EA + U Minnesota EA. If the ED admits you, you withdraw the rest.","flag"); h6(R,44); R+=1
blk(R,"   Lane C — All non-binding: MIT EA + UChicago EA + RIT EA + U Minnesota EA. No SCEA, no ED. Keeps every option open.","flag"); h6(R,30); R+=1
blk(R,"Key tension: going Yale-early (Lane A) means NO early MIT and NO early UChicago — two of Toby's three Ambitious schools. "
      "If MIT and UChicago early matter more than Yale early, Lane B or C is the stronger play.","body"); h6(R,42); R+=2

blk(R,"1a. Toby's special option: UChicago SSEN (ED0) — resolves BEFORE Nov 1","h"); R+=1
blk(R,"Because Toby completes UChicago's Intensive Japanese pre-college program this summer, he's eligible for Summer Student "
      "Early Notification (SSEN / ED0): apply Sept 1–Oct 15 and hear back before Nov 1 — before any other early app. Lighter-weight, "
      "too: reuse the Summer Session teacher recs, and UChicago will review without a counselor letter if one isn't ready.","body"); h6(R,58); R+=1
blk(R,"   Upside — If UChicago is his #1, this is the strongest form of Lane B: an early, binding answer with less paperwork and a "
      "calmer fall. If admitted, he enrolls and withdraws the rest.","flag"); h6(R,44); R+=1
blk(R,"   Catch — SSEN is a BINDING private ED, so it rules out Yale's SCEA. If deferred, he's released and can still pivot to MIT EA "
      "(non-restrictive) and Regular Decision everywhere — just not Yale-early. Confirm edge cases with his counselor.","flag"); h6(R,44); R+=2

blk(R,"2. Binding vs. non-binding — the quick key (this shortlist)","h"); R+=1
blk(R,"BINDING (enroll if admitted; only ONE allowed): UChicago ED I/II (and SSEN), UPenn, Columbia, Swarthmore, Vassar, Wesleyan, RIT.","body"); h6(R,30); R+=1
blk(R,"NON-BINDING: MIT EA, UChicago EA, Yale SCEA (restrictive but non-binding), RIT EA, University of Minnesota EA I/II, and all Regular Decision.","body"); h6(R,30); R+=2

blk(R,"3. Testing — Toby is covered","h"); R+=1
blk(R,"On this shortlist, SAT/ACT are REQUIRED at MIT, Yale, and UPenn. Everyone else is test-optional this cycle (Columbia optional "
      "now but required next year). Submit his 1590 SAT / 38 ACT everywhere — required at three, and helpful at the rest.","body"); h6(R,44); R+=2

blk(R,"4. Recommenders — one set covers the list","h"); R+=1
blk(R,"Counselor + one math/science teacher + one humanities/language teacher satisfies MIT's split and every 2-teacher school "
      "(Yale, UChicago, UPenn, Columbia, Swarthmore, Vassar, Wesleyan). RIT needs just one; University of Minnesota needs none.","body"); h6(R,44); R+=2

blk(R,"5. Distinctive requirements not to miss","h"); R+=1
blk(R,"• MIT: separate application system + a 'February Updates & Notes' form for midyear grades.\n"
      "• UChicago: no CSS Profile (FAFSA + its own worksheet); SSEN option above.\n"
      "• UPenn: you apply to one of four schools (A&S, Wharton, Engineering, Nursing), each with its own essay.\n"
      "• Columbia: the list question + five short essays; a 'Why major' that differs for Columbia College vs. Engineering.\n"
      "• University of Minnesota: apply via Common App OR the Golden Gopher app; no recommendations; one short 'which college/major' essay; self-reported grades.\n"
      "• RIT: ED I/II are binding; if submitting the ACT with a STEM focus, add the optional Science section.","body"); h6(R,116); R+=2

blk(R,"6. Yield protection — worth a note","h"); R+=1
blk(R,"The selective LACs on this list — Swarthmore, Vassar, and Wesleyan — sometimes waitlist an obviously elite-bound 1590 applicant "
      "they doubt will enrol. Demonstrated interest (a visit, a specific 'why us') matters most exactly there.","body"); h6(R,44); R+=2

blk(R,"Sources: official MIT, Yale, UChicago, UPenn, Columbia, Swarthmore, Vassar, Wesleyan, RIT, and University of Minnesota "
      "admissions & financial-aid pages. This is a working subset chosen for planning — all 21 researched schools remain in the "
      "Application Tracker. Confirm any date marked * once 2026–27 dates post.","body")
ws4.cell(row=R,column=2).font=Font(name=FONT,italic=True,size=8,color="808080"); h6(R,44)

# =========================================================
# TAB 6 — DECISION MODEL (editable utilities & likelihoods -> EV & PV)
# =========================================================
ws5=wb.create_sheet("Decision Model")
YEL="FFF2CC"; G1="F2F2F2"
def put(row,col,val,bold=False,fill=None,align="left",size=10,color="000000",wrap=False,num=None,italic=False,bd=False):
    c=ws5.cell(row=row,column=col,value=val)
    c.font=Font(name=FONT,bold=bold,size=size,color=color,italic=italic)
    if fill:c.fill=PatternFill("solid",fgColor=fill)
    c.alignment=Alignment(horizontal=align,vertical="center",wrap_text=wrap)
    if num:c.number_format=num
    if bd:c.border=border
    return c

title_row(ws5,"Decision Model — early-lane strategy: editable utilities & likelihoods → EV / PV","J")
ws5.merge_cells("A2:J2")
put(2,1,"Edit only the YELLOW cells (utilities 0–100 and admission probabilities 0–1). EV and PV recompute automatically. Higher is better. Keep utilities distinct and probabilities ≤ 0.98.",italic=True,size=9,color="595959",wrap=True)
ws5.row_dimensions[2].height=26

# --- quick decision tree ---
put(4,1,"Quick read",bold=True,size=11,color=NAVY)
tree=[
"• If UChicago is Toby's clear #1  →  SSEN (Strategy B): earliest, lightest, binding — and a deferral still leaves MIT EA + everything Regular.",
"• If Yale is the clear #1  →  Yale SCEA (Strategy A): but this forecloses early MIT and early UChicago.",
"• If keeping options open & comparing aid matters most  →  All non-binding EA (Strategy C).",
"• A binding ED elsewhere (D / E) trades upside for certainty — the model prices that out.",
"• Before ANY binding choice, run the school's Net Price Calculator so you're not committing blind on aid.",
]
rr=5
for t in tree:
    ws5.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=10)
    put(rr,1,t,size=9,wrap=True); ws5.row_dimensions[rr].height=15; rr+=1

# --- RESULTS ---
put(11,1,"RESULTS — ranked by PV (edit inputs below to update)",bold=True,size=11,color=WHITE,fill=NAVY)
for cc in range(2,11): put(11,cc,"",fill=NAVY)
ws5.merge_cells("A11:J11")
put(11,1,"RESULTS — each early-lane strategy (edit inputs below; everything recomputes)",bold=True,size=11,color=WHITE,fill=NAVY)
for j,h in enumerate(["Strategy","What it means","EV","PV","",""],1):
    put(12,j,h,bold=True,fill=BLUE,color=WHITE,align="center" if j>2 else "left",bd=True)
put(12,3,"EV (0–100)",bold=True,fill=BLUE,color=WHITE,align="center",bd=True)
put(12,4,"PV (net)",bold=True,fill=BLUE,color=WHITE,align="center",bd=True)
put(12,5,"Best?",bold=True,fill=BLUE,color=WHITE,align="center",bd=True)
res=[
("A — Yale SCEA (restrictive EA)","Yale early, non-binding. Rules out early MIT & UChicago; everything else Regular.","=C52","=C13+C35*C23"),
("B — UChicago SSEN (ED0)","Binding, resolves before Nov 1. A deferral still leaves MIT EA + all Regular.","=C34*B24+(1-C34)*G52","=C14+C35*C34-C36*C34"),
("C — All non-binding EA","MIT EA + UChicago EA + Minnesota EA; everything else Regular. Preserves aid comparison.","=E52","=C15+C35*(1-(1-C22)*(1-C24))"),
("D — UPenn ED (a reach ED)","Binding Penn ED at Nov 1. If denied, no MIT EA left — all Regular.","=C25*B25+(1-C25)*I52","=C16+C35*C25-C36*C25"),
("E — Vassar ED (lock a likely)","Binding Vassar ED. High certainty, but commits away the reaches.","=C28*B28+(1-C28)*I52","=C17+C35*C28-C36*C28"),
]
r=13
for name,mean,ev,pv in res:
    put(r,1,name,bold=True,bd=True)
    put(r,2,mean,size=9,wrap=True,bd=True)
    put(r,3,ev,align="center",num="0.0",bd=True)
    put(r,4,pv,align="center",num="0.0",bold=True,bd=True,fill=GREEN if False else None)
    put(r,5,f'=IF(D{r}=MAX($D$13:$D$17),"◄ best","")',align="center",bold=True,color="38761D",bd=True)
    ws5.row_dimensions[r].height=30
    r+=1
put(18,1,"EV = expected value of where Toby enrolls (utility of his best admit, probability-weighted).  PV = EV + certainty/less-work bonus when the early round resolves − aid-comparison penalty when he's bound before seeing aid.",italic=True,size=8,color="808080",wrap=True)
ws5.merge_cells("A18:J18"); ws5.row_dimensions[18].height=26

# --- INPUTS ---
put(20,1,"INPUTS — edit the yellow cells",bold=True,size=11,color=WHITE,fill=NAVY)
ws5.merge_cells("A20:J20")
put(20,1,"INPUTS — edit the yellow cells (utilities & probabilities)",bold=True,size=11,color=WHITE,fill=NAVY)
for j,h in enumerate(["School","Utility 0–100","P(admit) early","P(admit) Regular"],1):
    put(21,j,h,bold=True,fill=BLUE,color=WHITE,align="center" if j>1 else "left",bd=True)
put(21,6,"What \"early\" means here",bold=True,fill=BLUE,color=WHITE,align="left",bd=True)
schools=[("MIT",100,0.10,0.08,"EA (non-binding)"),
("Yale",98,0.12,0.06,"SCEA (restrictive, non-binding)"),
("University of Chicago",92,0.20,0.12,"EA / ED I  (SSEN handled below)"),
("UPenn",80,0.22,0.09,"ED (binding)"),
("Columbia",82,0.18,0.08,"ED (binding)"),
("Swarthmore",78,0.35,0.15,"ED (binding)"),
("Vassar",68,0.55,0.35,"ED (binding)"),
("Wesleyan",66,0.55,0.35,"ED (binding)"),
("RIT",60,0.90,0.85,"EA / ED (EA non-binding)"),
("University of Minnesota",55,0.97,0.95,"EA (non-binding public)")]
r=22
for nm,u,pe,pr,note in schools:
    put(r,1,nm,bold=True,bd=True,fill=GREEN2 if nm in("MIT","Yale","University of Chicago","UPenn","Columbia","Swarthmore","Vassar","Wesleyan","RIT","University of Minnesota") else None)
    put(r,2,u,align="center",fill=YEL,num="0",bd=True)
    put(r,3,pe,align="center",fill=YEL,num="0.00",bd=True)
    put(r,4,pr,align="center",fill=YEL,num="0.00",bd=True)
    put(r,6,note,size=9,bd=True)
    r+=1
put(34,1,"UChicago SSEN (ED0) admit probability",bd=True); put(34,3,0.35,align="center",fill=YEL,num="0.00",bd=True)
put(35,1,"Certainty / less-work bonus (utility pts)",bd=True); put(35,3,4,align="center",fill=YEL,num="0",bd=True)
put(36,1,"Aid-comparison penalty if bound (utility pts)",bd=True); put(36,3,6,align="center",fill=YEL,num="0",bd=True)

# --- ENGINE ---
put(38,1,"COMPUTATION ENGINE (safe to ignore — this is what feeds the results)",bold=True,size=9,color="808080")
eng_hdr=["School","U","A: p","A: enroll","C: p","C: enroll","B-fb: p","B-fb: enroll","RD: p","RD: enroll"]
for j,h in enumerate(eng_hdr,1): put(40,j,h,bold=True,size=8,fill=G1,align="center",bd=True)
# per-school probability mappings (row offset: inputs 22..31 -> engine 41..50)
Cmap=["D22","C23","D24","D25","D26","D27","D28","D29","D30","C31"]   # V1 Yale SCEA
Emap=["C22","D23","C24","D25","D26","D27","D28","D29","C30","C31"]   # V2 all-EA
Gmap=["C22","D23","D24","D25","D26","D27","D28","D29","C30","C31"]   # V3 SSEN fallback
Imap=["D22","D23","D24","D25","D26","D27","D28","D29","D30","D31"]   # RD
for i in range(10):
    r=41+i; sref=22+i
    put(r,1,f"=A{sref}",size=8,bd=True)
    put(r,2,f"=B{sref}",size=8,align="center",num="0",bd=True)
    put(r,3,f"={Cmap[i]}",size=8,align="center",num="0.00",bd=True)
    put(r,4,f"=C{r}*EXP(SUMPRODUCT(($B$41:$B$50>$B{r})*LN(1-$C$41:$C$50)))",size=8,align="center",num="0.000",bd=True)
    put(r,5,f"={Emap[i]}",size=8,align="center",num="0.00",bd=True)
    put(r,6,f"=E{r}*EXP(SUMPRODUCT(($B$41:$B$50>$B{r})*LN(1-$E$41:$E$50)))",size=8,align="center",num="0.000",bd=True)
    put(r,7,f"={Gmap[i]}",size=8,align="center",num="0.00",bd=True)
    put(r,8,f"=G{r}*EXP(SUMPRODUCT(($B$41:$B$50>$B{r})*LN(1-$G$41:$G$50)))",size=8,align="center",num="0.000",bd=True)
    put(r,9,f"={Imap[i]}",size=8,align="center",num="0.00",bd=True)
    put(r,10,f"=I{r}*EXP(SUMPRODUCT(($B$41:$B$50>$B{r})*LN(1-$I$41:$I$50)))",size=8,align="center",num="0.000",bd=True)
put(52,1,"Best-admit EV of each vector →",size=8,bold=True,align="right")
put(52,3,"=SUMPRODUCT($B$41:$B$50,$D$41:$D$50)",size=8,align="center",num="0.0",bd=True,fill=G1)
put(52,5,"=SUMPRODUCT($B$41:$B$50,$F$41:$F$50)",size=8,align="center",num="0.0",bd=True,fill=G1)
put(52,7,"=SUMPRODUCT($B$41:$B$50,$H$41:$H$50)",size=8,align="center",num="0.0",bd=True,fill=G1)
put(52,9,"=SUMPRODUCT($B$41:$B$50,$J$41:$J$50)",size=8,align="center",num="0.0",bd=True,fill=G1)
for k,v in {"A":26,"B":40,"C":12,"D":11,"E":13,"F":24,"G":12,"H":12,"I":12,"J":12}.items(): ws5.column_dimensions[k].width=v

os.makedirs("output", exist_ok=True)
wb.save(os.path.join("output","Toby_College_Tracker.xlsx"))
print("saved")
